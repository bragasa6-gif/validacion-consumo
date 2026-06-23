import io
from datetime import date

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials

st.set_page_config(
    page_title="Validación de Consumo",
    page_icon="🦐",
    layout="centered"
)

BASE_FILE = "tabla_base.xlsx"
SHEET_ID = st.secrets["SHEET_ID"]

st.markdown(
    """
    <style>
    .main-title {
        text-align: center;
        font-size: 38px;
        font-weight: 800;
        margin-bottom: 0px;
    }
    .subtitle {
        text-align: center;
        color: #6b7280;
        font-size: 16px;
        margin-bottom: 25px;
    }
    .card {
        background: #ffffff;
        padding: 18px;
        border-radius: 18px;
        box-shadow: 0 4px 18px rgba(0,0,0,0.08);
        margin-bottom: 14px;
        border: 1px solid #eef2f7;
    }
    .metric-label {
        color: #6b7280;
        font-size: 14px;
        margin-bottom: 4px;
    }
    .metric-value {
        font-size: 30px;
        font-weight: 800;
        color: #111827;
    }
    .semaforo {
        text-align: center;
        font-size: 42px;
        font-weight: 900;
        padding: 24px;
        border-radius: 24px;
        margin-bottom: 20px;
    }
    .verde {
        background: #dcfce7;
        color: #166534;
    }
    .amarillo {
        background: #fef9c3;
        color: #854d0e;
    }
    .rojo {
        background: #fee2e2;
        color: #991b1b;
    }
    .gris {
        background: #f3f4f6;
        color: #374151;
    }
    </style>
    """,
    unsafe_allow_html=True
)


def to_float(value):
    if value is None:
        return np.nan

    if isinstance(value, str):
        text = value.strip()
        text = text.replace("%", "")
        text = text.replace(".", "")
        text = text.replace(",", ".")

        if text in ("", "-", "#DIV/0!", "#VALUE!", "#N/A", "#REF!", "#NAME?", "#NUM!", "#NULL!"):
            return np.nan

        try:
            return float(text)
        except ValueError:
            return np.nan

    try:
        return float(value)
    except Exception:
        return np.nan


def normalize_bw(value):
    value = to_float(value)

    if pd.isna(value):
        return np.nan

    return value / 100 if value > 1 else value


@st.cache_data(show_spinner=False)
def load_tables_from_bytes(file_bytes):
    wb = load_workbook(
        io.BytesIO(file_bytes),
        data_only=True,
        read_only=True
    )

    sheet_name = None

    for name in wb.sheetnames:
        if name.strip().lower() == "tabla":
            sheet_name = name
            break

    if sheet_name is None:
        raise ValueError("No encontré la hoja Tabla en el Excel.")

    ws = wb[sheet_name]

    # Hoja Tabla:
    # A = Peso
    # B = Normal Frío
    # C = Normal Calor
    normal_rows = []

    for row in ws.iter_rows(
        min_row=4,
        max_row=200,
        min_col=1,
        max_col=3,
        values_only=True
    ):
        peso, frio, calor = row
        peso = to_float(peso)

        if pd.isna(peso):
            continue

        normal_rows.append(
            {
                "peso": peso,
                "NormalFrio": normalize_bw(frio),
                "NormalCalor": normalize_bw(calor),
            }
        )

    # Hoja Tabla:
    # F = Peso
    # G = Min Calor
    # H = Min Frío
    min_rows = []

    for row in ws.iter_rows(
        min_row=3,
        max_row=600,
        min_col=6,
        max_col=8,
        values_only=True
    ):
        peso, calor, frio = row
        peso = to_float(peso)

        if pd.isna(peso):
            continue

        min_rows.append(
            {
                "peso": peso,
                "MinCalor": normalize_bw(calor),
                "MinFrio": normalize_bw(frio),
            }
        )

    normal = pd.DataFrame(normal_rows).sort_values("peso").reset_index(drop=True)
    minimo = pd.DataFrame(min_rows).sort_values("peso").reset_index(drop=True)

    if normal.empty or minimo.empty:
        raise ValueError("No pude leer las tablas A:C y F:H de la hoja Tabla.")

    return normal, minimo


def vlookup_approx(table, peso, column):
    valid = table[table["peso"] <= peso]

    if valid.empty:
        row = table.iloc[0]
    else:
        row = valid.iloc[-1]

    return float(row[column]), float(row["peso"])


def safe_div(a, b):
    if b is None or b == 0 or pd.isna(b):
        return np.nan

    return a / b


def fmt_num(value, decimals=0):
    if pd.isna(value):
        return "-"
    return f"{value:,.{decimals}f}"


def metric_card(label, value):
    st.markdown(
        f"""
        <div class="card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{value}</div>
        </div>
        """,
        unsafe_allow_html=True
    )


def semaforo_html(estado):
    if "Verde" in estado:
        css = "verde"
    elif "Amarillo" in estado:
        css = "amarillo"
    elif "Rojo" in estado:
        css = "rojo"
    else:
        css = "gris"

    st.markdown(
        f"""
        <div class="semaforo {css}">
            {estado}
        </div>
        """,
        unsafe_allow_html=True
    )


def estado_por_diferencia(pct):
    if pd.isna(pct):
        return "⚪ Sin referencia"

    if abs(pct) <= 5:
        return "🟢 Verde"

    if abs(pct) <= 10:
        return "🟡 Amarillo"

    return "🔴 Rojo"

def guardar_google_sheets(registro):
    try:

        scope = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]

        creds = Credentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]),
            scopes=scope
        )

        client = gspread.authorize(creds)

        sheet = client.open_by_key(SHEET_ID).sheet1

        fila = list(registro.values())

        sheet.append_row(fila)

    except Exception as e:
        st.error(f"Error guardando en Google Sheets: {e}")

def calcular_bw(epoca, peso_actual, normal_table, min_table):
    # CORREGIDO:
    # Normal: B = Frío, C = Calor
    # Mínima: G = Min Calor, H = Min Frío
    normal_col = "NormalCalor" if epoca == "Calor" else "NormalFrio"
    min_col = "MinCalor" if epoca == "Calor" else "MinFrio"

    bw_sug, peso_ref_normal = vlookup_approx(
        normal_table,
        peso_actual,
        normal_col
    )

    bw_min, peso_ref_min = vlookup_approx(
        min_table,
        peso_actual,
        min_col
    )

    return bw_sug, bw_min, peso_ref_normal, peso_ref_min


st.markdown("<div class='main-title'>🦐 Validación de Consumo</div>", unsafe_allow_html=True)
st.markdown("<div class='subtitle'>Cálculo rápido de población por consumo desde el celular</div>", unsafe_allow_html=True)


with st.sidebar:
    st.header("Tabla base")

    uploaded_base = st.file_uploader(
        "Subir Excel base",
        type=["xlsx"],
        help="Debe contener la hoja Tabla"
    )

    usar_incluida = st.toggle(
        "Usar tabla_base.xlsx incluida",
        value=True
    )

    if uploaded_base is not None:
        base_bytes = uploaded_base.getvalue()
        base_label = uploaded_base.name

    elif usar_incluida:
        with open(BASE_FILE, "rb") as f:
            base_bytes = f.read()

        base_label = BASE_FILE

    else:
        st.info("Sube el Excel base para iniciar.")
        st.stop()


try:
    normal_table, min_table = load_tables_from_bytes(base_bytes)

    st.sidebar.success("Tabla cargada")
    st.sidebar.caption(base_label)

except Exception as exc:
    st.error(f"No pude leer la hoja Tabla: {exc}")
    st.stop()


with st.expander("Ver tablas de referencia", expanded=False):
    st.write("Tabla Normal: A = Peso, B = Normal Frío, C = Normal Calor")
    st.dataframe(normal_table, hide_index=True)

    st.write("Tabla Mínima: F = Peso, G = Min Calor, H = Min Frío")
    st.dataframe(min_table, hide_index=True)


escenario = st.radio(
    "¿Qué quieres calcular?",
    [
        "1. Validar población reportada",
        "2. Estimar población por consumo",
        "3. Estimar densidad por consumo diario"
    ]
)


if "historial" not in st.session_state:
    st.session_state.historial = []


if escenario == "1. Validar población reportada":
    st.subheader("Escenario 1: Validar población reportada")

    with st.form("form_validar"):
        fecha = st.date_input("Fecha", value=date.today())

        psc = st.text_input("Piscina / PSC", value="")

        marca = st.text_input("Marca alimento", value="")

        area_ha = st.number_input(
            "Área del estanque (ha)",
            min_value=0.0,
            value=0.0,
            step=0.10,
            format="%.2f"
        )

        densidad_siembra_m2 = st.number_input(
            "Densidad inicial de siembra (cam/m²)",
            min_value=0.0,
            value=0.0,
            step=0.10,
            format="%.2f"
        )

        peso_actual = st.number_input(
            "Peso actual (g)",
            min_value=0.01,
            value=3.50,
            step=0.10,
            format="%.2f"
        )

        kg_alimento_actual = st.number_input(
            "Kg alimento semanal actual",
            min_value=0.0,
            value=0.0,
            step=1.0
        )

        dias_alimentados = st.number_input(
            "Días alimentados",
            min_value=1.0,
            value=7.0,
            step=1.0
        )

        animales_vivos = st.number_input(
            "Animales vivos reportados en campo",
            min_value=0.0,
            value=0.0,
            step=1000.0
        )

        epoca = st.radio(
            "Época",
            ["Calor", "Frío"],
            horizontal=True
        )

        calcular = st.form_submit_button("Calcular validación")

    if calcular:
        bw_sug, bw_min, peso_ref_normal, peso_ref_min = calcular_bw(
            epoca,
            peso_actual,
            normal_table,
            min_table
        )

        area_m2 = area_ha * 10000
        animales_sembrados = area_m2 * densidad_siembra_m2

        biomasa_reportada = animales_vivos * peso_actual / 1000

        kg_sugerido = biomasa_reportada * bw_sug
        kg_minimo = biomasa_reportada * bw_min

        kg_diario_actual = safe_div(
            kg_alimento_actual,
            dias_alimentados
        )

        # Kg alimento sugerido ya es DIARIO.
        # No se divide para días alimentados.
        kg_diario_sugerido = kg_sugerido

        kg_100k_finca = safe_div(
            kg_diario_actual,
            safe_div(animales_vivos, 100000)
        )

        kg_100k_sugerido = safe_div(
            kg_diario_sugerido,
            safe_div(animales_vivos, 100000)
        )

        biomasa_actual = safe_div(
            kg_diario_actual,
            bw_sug
        )

        animales_ao = safe_div(
            biomasa_actual * 1000,
            peso_actual
        )

        biomasa_sugerida = safe_div(
            kg_sugerido,
            bw_sug
        )

        biomasa_minima = safe_div(
            kg_minimo,
            bw_min
        )

        animales_an = safe_div(
            biomasa_sugerida * 1000,
            peso_actual
        )

        animales_ap = safe_div(
            biomasa_minima * 1000,
            peso_actual
        )

        densidad_inicial_m2 = densidad_siembra_m2

        densidad_reportada_m2 = safe_div(
            animales_vivos,
            area_m2
        )

        densidad_actual_m2 = safe_div(
            animales_ao,
            area_m2
        )

        supervivencia_reportada = safe_div(
            animales_vivos,
            animales_sembrados
        ) * 100 if animales_sembrados else np.nan

        supervivencia_estimada = safe_div(
            animales_ao,
            animales_sembrados
        ) * 100 if animales_sembrados else np.nan

        dif_ao_al = animales_ao - animales_vivos if animales_vivos else np.nan
        dif_ao_al_pct = safe_div(dif_ao_al, animales_vivos) * 100 if animales_vivos else np.nan

        estado = estado_por_diferencia(dif_ao_al_pct)

        st.subheader("Resultado de validación")
        semaforo_html(estado)

        metric_card("Diferencia entre alimentación actual y animales reportados", fmt_num(dif_ao_al, 0))
        metric_card("Diferencia porcentual", f"{fmt_num(dif_ao_al_pct, 2)}%")

        st.subheader("Resumen principal")

        c1, c2 = st.columns(2)
        with c1:
            metric_card("Animales según alimentación actual", fmt_num(animales_ao, 0))
        with c2:
            metric_card("Animales vivos reportados", fmt_num(animales_vivos, 0))

        c3, c4 = st.columns(2)
        with c3:
            metric_card("Animales sembrados iniciales", fmt_num(animales_sembrados, 0))
        with c4:
            metric_card("Área del estanque", f"{fmt_num(area_ha, 2)} ha")

        st.subheader("Indicadores de alimento")

        c5, c6 = st.columns(2)
        with c5:
            metric_card("Kg/100.000 animales Finca", f"{fmt_num(kg_100k_finca, 2)} kg/día")
        with c6:
            metric_card("Kg/100.000 animales Sugerido", f"{fmt_num(kg_100k_sugerido, 2)} kg/día")

        st.subheader("Densidad y supervivencia")

        c5b, c6b = st.columns(2)
        with c5b:
            metric_card("Densidad inicial", f"{fmt_num(densidad_inicial_m2, 2)} cam/m²")
        with c6b:
            metric_card("Densidad actual estimada", f"{fmt_num(densidad_actual_m2, 2)} cam/m²")

        c7, c8 = st.columns(2)
        with c7:
            metric_card("Densidad reportada en campo", f"{fmt_num(densidad_reportada_m2, 2)} cam/m²")
        with c8:
            metric_card("Supervivencia estimada por consumo", f"{fmt_num(supervivencia_estimada, 2)}%")

        c9, c10 = st.columns(2)
        with c9:
            metric_card("Supervivencia reportada campo", f"{fmt_num(supervivencia_reportada, 2)}%")
        with c10:
            metric_card("Kg diario actual", f"{fmt_num(kg_diario_actual, 2)} kg/día")

        st.subheader("Detalle técnico")

        c11, c12 = st.columns(2)
        with c11:
            metric_card("BW sugerida", f"{fmt_num(bw_sug * 100, 2)}%")
        with c12:
            metric_card("BW mínima", f"{fmt_num(bw_min * 100, 2)}%")

        c13, c14 = st.columns(2)
        with c13:
            metric_card("Kg alimento sugerido diario", fmt_num(kg_sugerido, 0))
        with c14:
            metric_card("Kg mínimo diario", fmt_num(kg_minimo, 0))

        c15, c16 = st.columns(2)
        with c15:
            metric_card("Animales según alimentación sugerida", fmt_num(animales_an, 0))
        with c16:
            metric_card("Animales según tabla 2 Min", fmt_num(animales_ap, 0))

        st.caption(
            f"Referencia usada: BW sugerida con peso {peso_ref_normal} g | BW mínima con peso {peso_ref_min} g"
        )

        registro = {
            "Escenario": "Validar población reportada",
            "Fecha": fecha.isoformat(),
            "Piscina / PSC": psc,
            "Marca alimento": marca,
            "Área ha": area_ha,
            "Área m2": area_m2,
            "Densidad inicial cam/m2": densidad_siembra_m2,
            "Animales sembrados iniciales": animales_sembrados,
            "Peso actual": peso_actual,
            "Kg alimento semanal actual": kg_alimento_actual,
            "Días alimentados": dias_alimentados,
            "Kg diario actual": kg_diario_actual,
            "Kg diario sugerido": kg_diario_sugerido,
            "Kg/100k finca": kg_100k_finca,
            "Kg/100k sugerido": kg_100k_sugerido,
            "Animales vivos reportados": animales_vivos,
            "Densidad reportada cam/m2": densidad_reportada_m2,
            "Densidad actual estimada cam/m2": densidad_actual_m2,
            "Supervivencia reportada campo %": supervivencia_reportada,
            "Supervivencia estimada por consumo %": supervivencia_estimada,
            "Época": epoca,
            "BW sugerida %": bw_sug * 100,
            "BW mínima %": bw_min * 100,
            "Kg alimento sugerido diario": kg_sugerido,
            "Kg mínimo diario": kg_minimo,
            "Animales según alimentación sugerida": animales_an,
            "Animales según alimentación actual": animales_ao,
            "Animales según tabla 2 Min": animales_ap,
            "Diferencia AO vs AL": dif_ao_al,
            "Diferencia AO vs AL %": dif_ao_al_pct,
            "Semáforo": estado,
        }

        st.session_state.historial.append(registro)
guardar_google_sheets(registro)

if escenario == "2. Estimar población por consumo":
    st.subheader("Escenario 2: Estimar población por consumo")

    with st.form("form_estimar"):
        fecha = st.date_input("Fecha", value=date.today())

        psc = st.text_input("Piscina / PSC", value="")

        marca = st.text_input("Marca alimento", value="")

        area_ha = st.number_input(
            "Área del estanque (ha)",
            min_value=0.0,
            value=0.0,
            step=0.10,
            format="%.2f"
        )

        densidad_siembra_m2 = st.number_input(
            "Densidad inicial de siembra (cam/m²)",
            min_value=0.0,
            value=0.0,
            step=0.10,
            format="%.2f"
        )

        peso_actual = st.number_input(
            "Peso actual (g)",
            min_value=0.01,
            value=3.50,
            step=0.10,
            format="%.2f"
        )

        kg_alimento_actual = st.number_input(
            "Kg alimento semanal actual",
            min_value=0.0,
            value=0.0,
            step=1.0
        )

        dias_alimentados = st.number_input(
            "Días alimentados",
            min_value=1.0,
            value=7.0,
            step=1.0
        )

        epoca = st.radio(
            "Época",
            ["Calor", "Frío"],
            horizontal=True
        )

        calcular = st.form_submit_button("Estimar población")

    if calcular:
        bw_sug, bw_min, peso_ref_normal, peso_ref_min = calcular_bw(
            epoca,
            peso_actual,
            normal_table,
            min_table
        )

        area_m2 = area_ha * 10000
        animales_sembrados = area_m2 * densidad_siembra_m2

        kg_diario_actual = safe_div(
            kg_alimento_actual,
            dias_alimentados
        )

        biomasa_actual = safe_div(
            kg_diario_actual,
            bw_sug
        )

        animales_estimados = safe_div(
            biomasa_actual * 1000,
            peso_actual
        )

        densidad_actual_m2 = safe_div(
            animales_estimados,
            area_m2
        )

        kg_100k_estimado = safe_div(
            kg_diario_actual,
            safe_div(animales_estimados, 100000)
        )

        supervivencia_estimada = safe_div(
            animales_estimados,
            animales_sembrados
        ) * 100 if animales_sembrados else np.nan

        dif_animales = animales_estimados - animales_sembrados if animales_sembrados else np.nan
        dif_animales_pct = safe_div(dif_animales, animales_sembrados) * 100 if animales_sembrados else np.nan

        estado = estado_por_diferencia(dif_animales_pct)

        st.subheader("Resultado de estimación")
        semaforo_html(estado)

        st.subheader("Resumen principal")

        c1, c2 = st.columns(2)
        with c1:
            metric_card("Animales estimados por consumo", fmt_num(animales_estimados, 0))
        with c2:
            metric_card("Animales sembrados iniciales", fmt_num(animales_sembrados, 0))

        c3, c4 = st.columns(2)
        with c3:
            metric_card("Densidad actual en agua", f"{fmt_num(densidad_actual_m2, 2)} cam/m²")
        with c4:
            metric_card("Densidad inicial de siembra", f"{fmt_num(densidad_siembra_m2, 2)} cam/m²")

        c5, c6 = st.columns(2)
        with c5:
            metric_card("Supervivencia estimada", f"{fmt_num(supervivencia_estimada, 2)}%")
        with c6:
            metric_card("Área del estanque", f"{fmt_num(area_ha, 2)} ha")

        st.subheader("Indicadores de alimento")
        metric_card("Kg/100.000 animales estimado", f"{fmt_num(kg_100k_estimado, 2)} kg/día")

        st.subheader("Detalle técnico")

        c7, c8 = st.columns(2)
        with c7:
            metric_card("BW sugerida", f"{fmt_num(bw_sug * 100, 2)}%")
        with c8:
            metric_card("BW mínima", f"{fmt_num(bw_min * 100, 2)}%")

        c9, c10 = st.columns(2)
        with c9:
            metric_card("Kg diario actual", f"{fmt_num(kg_diario_actual, 2)} kg/día")
        with c10:
            metric_card("Biomasa estimada", f"{fmt_num(biomasa_actual, 2)} kg")

        st.caption(
            f"Referencia usada: BW sugerida con peso {peso_ref_normal} g | BW mínima con peso {peso_ref_min} g"
        )

        registro = {
            "Escenario": "Estimar población por consumo",
            "Fecha": fecha.isoformat(),
            "Piscina / PSC": psc,
            "Marca alimento": marca,
            "Área ha": area_ha,
            "Área m2": area_m2,
            "Densidad inicial cam/m2": densidad_siembra_m2,
            "Animales sembrados iniciales": animales_sembrados,
            "Peso actual": peso_actual,
            "Kg alimento semanal actual": kg_alimento_actual,
            "Días alimentados": dias_alimentados,
            "Kg diario actual": kg_diario_actual,
            "Kg/100k estimado": kg_100k_estimado,
            "Época": epoca,
            "BW sugerida %": bw_sug * 100,
            "BW mínima %": bw_min * 100,
            "Biomasa estimada kg": biomasa_actual,
            "Animales estimados por consumo": animales_estimados,
            "Densidad actual cam/m2": densidad_actual_m2,
            "Supervivencia estimada %": supervivencia_estimada,
            "Diferencia estimados vs sembrados": dif_animales,
            "Diferencia estimados vs sembrados %": dif_animales_pct,
            "Semáforo": estado,
        }

        st.session_state.historial.append(registro)
guardar_google_sheets(registro)


if escenario == "3. Estimar densidad por consumo diario":
    st.subheader("Escenario 3: Estimar densidad por consumo diario")

    with st.form("form_densidad_consumo"):
        fecha = st.date_input("Fecha", value=date.today())
        psc = st.text_input("Piscina / PSC", value="")
        marca = st.text_input("Marca alimento", value="")

        area_ha = st.number_input(
            "Área del estanque (ha)",
            min_value=0.0,
            value=0.0,
            step=0.10,
            format="%.2f"
        )

        peso_actual = st.number_input(
            "Peso actual (g)",
            min_value=0.01,
            value=3.50,
            step=0.10,
            format="%.2f"
        )

        kg_diario_total = st.number_input(
            "Kg diario total del estanque",
            min_value=0.0,
            value=0.0,
            step=1.0
        )

        epoca = st.radio(
            "Época",
            ["Calor", "Frío"],
            horizontal=True
        )

        calcular = st.form_submit_button("Estimar densidad")

    if calcular:
        bw_sug, bw_min, peso_ref_normal, peso_ref_min = calcular_bw(
            epoca,
            peso_actual,
            normal_table,
            min_table
        )

        area_m2 = area_ha * 10000

        biomasa_estimada = safe_div(
            kg_diario_total,
            bw_sug
        )

        animales_estimados = safe_div(
            biomasa_estimada * 1000,
            peso_actual
        )

        densidad_actual_m2 = safe_div(
            animales_estimados,
            area_m2
        )

        kg_100k_estimado = safe_div(
            kg_diario_total,
            safe_div(animales_estimados, 100000)
        )

        st.subheader("Resultado de estimación")

        c1, c2 = st.columns(2)
        with c1:
            metric_card("Densidad actual en agua", f"{fmt_num(densidad_actual_m2, 2)} cam/m²")
        with c2:
            metric_card("Biomasa estimada", f"{fmt_num(biomasa_estimada, 2)} kg")

        c3, c4 = st.columns(2)
        with c3:
            metric_card("Animales estimados", fmt_num(animales_estimados, 0))
        with c4:
            metric_card("Área del estanque", f"{fmt_num(area_ha, 2)} ha")

        st.subheader("Detalle técnico")

        c5, c6 = st.columns(2)
        with c5:
            metric_card("BW sugerida usada", f"{fmt_num(bw_sug * 100, 2)}%")
        with c6:
            metric_card("Kg diario total", f"{fmt_num(kg_diario_total, 2)} kg/día")

        metric_card("Kg/100.000 animales estimado", f"{fmt_num(kg_100k_estimado, 2)} kg/día")

        st.caption(
            f"Referencia usada: BW sugerida con peso {peso_ref_normal} g | BW mínima con peso {peso_ref_min} g"
        )

        registro = {
            "Escenario": "Estimar densidad por consumo diario",
            "Fecha": fecha.isoformat(),
            "Piscina / PSC": psc,
            "Marca alimento": marca,
            "Área ha": area_ha,
            "Área m2": area_m2,
            "Peso actual": peso_actual,
            "Kg diario total": kg_diario_total,
            "Época": epoca,
            "BW sugerida %": bw_sug * 100,
            "BW mínima %": bw_min * 100,
            "Biomasa estimada kg": biomasa_estimada,
            "Animales estimados": animales_estimados,
            "Densidad actual cam/m2": densidad_actual_m2,
            "Kg/100k estimado": kg_100k_estimado,
        }

        st.session_state.historial.append(registro)
guardar_google_sheets(registro)

st.subheader("Historial de esta sesión")

if st.session_state.historial:
    hist = pd.DataFrame(st.session_state.historial)

    st.dataframe(
        hist,
        hide_index=True
    )

    st.download_button(
        "Descargar historial",
        data=hist.to_csv(index=False).encode("utf-8-sig"),
        file_name="historial_validacion_consumo.csv",
        mime="text/csv",
    )

else:
    st.info("Aún no hay registros calculados.")
